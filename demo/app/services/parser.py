from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import csv


@dataclass
class ParsedBlock:
    section_path: str
    block_text: str
    block_type: str = "paragraph"


class DocumentParser:
    def parse_file(self, file_path: Path) -> list[ParsedBlock]:
        suffix = file_path.suffix.lower()
        if suffix == ".txt":
            raw_text = self._read_text(file_path)
            return self.split_text(raw_text)
        if suffix == ".md":
            raw_text = self._read_text(file_path)
            return self._parse_markdown(raw_text)
        if suffix == ".pdf":
            return self._parse_pdf(file_path)
        if suffix == ".docx":
            return self._parse_docx(file_path)
        if suffix in {".xlsx", ".xls"}:
            return self._parse_excel(file_path)
        if suffix == ".csv":
            return self._parse_csv(file_path)
        raw_text = self._read_text(file_path)
        return self.split_text(raw_text)

    def split_text(self, raw_text: str) -> list[ParsedBlock]:
        normalized = raw_text.replace("\r\n", "\n").replace("\r", "\n")
        paragraphs = [line.strip() for line in normalized.split("\n") if line.strip()]
        return [ParsedBlock(section_path=f"paragraph_{index + 1}", block_text=paragraph) for index, paragraph in enumerate(paragraphs)]

    def _read_text(self, file_path: Path) -> str:
        return file_path.read_text(encoding="utf-8", errors="ignore")

    def _parse_markdown(self, raw_text: str) -> list[ParsedBlock]:
        lines = [line.strip() for line in raw_text.splitlines() if line.strip()]
        return [ParsedBlock(section_path=f"md_line_{i + 1}", block_text=line) for i, line in enumerate(lines)]

    def _parse_pdf(self, file_path: Path) -> list[ParsedBlock]:
        try:
            import fitz
        except ImportError:
            return self.split_text(self._read_text(file_path))

        doc = fitz.open(file_path)
        blocks: list[ParsedBlock] = []
        for page_index in range(len(doc)):
            page = doc.load_page(page_index)
            text = page.get_text()
            paragraphs = [line.strip() for line in text.splitlines() if line.strip()]
            for i, paragraph in enumerate(paragraphs):
                blocks.append(ParsedBlock(section_path=f"page_{page_index + 1}_para_{i + 1}", block_text=paragraph))
        return blocks

    def _parse_docx(self, file_path: Path) -> list[ParsedBlock]:
        try:
            from docx import Document
        except ImportError:
            return self.split_text(self._read_text(file_path))

        doc = Document(file_path)
        blocks: list[ParsedBlock] = []
        for i, paragraph in enumerate(doc.paragraphs):
            text = paragraph.text.strip()
            if not text:
                continue
            blocks.append(ParsedBlock(section_path=f"docx_para_{i + 1}", block_text=text))
        return blocks

    def _parse_excel(self, file_path: Path) -> list[ParsedBlock]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return self.split_text(self._read_text(file_path))

        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        blocks: list[ParsedBlock] = []
        for sheet in workbook.worksheets:
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if not values:
                    continue
                blocks.append(ParsedBlock(section_path=f"{sheet.title}_row_{row_index}", block_text=" | ".join(values), block_type="table_row"))
        return blocks

    def _parse_csv(self, file_path: Path) -> list[ParsedBlock]:
        blocks: list[ParsedBlock] = []
        with file_path.open("r", encoding="utf-8", errors="ignore") as handle:
            reader = csv.reader(handle)
            for row_index, row in enumerate(reader, start=1):
                values = [cell.strip() for cell in row if cell.strip()]
                if not values:
                    continue
                blocks.append(ParsedBlock(section_path=f"csv_row_{row_index}", block_text=" | ".join(values), block_type="table_row"))
        return blocks
